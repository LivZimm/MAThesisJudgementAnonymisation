from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- PATHS ---
INPUT_XLSX = r"X.xlsx"
OUTPUT_XLSX = r"Y.xlsx"

# --- BEHAVIOUR ---
SEARCH_FIRST_N_CHARS = 25000
NEARBY_WINDOW = 500  # enlarged: OCR linebreaks etc.

# thresholds for date-only linking
DATE_MATCH_SIM_THRESHOLD = 0.10          # TFIDF (if available)
HARD_MARKER_THRESHOLD = 0.20             # hard-marker Jaccard
DATE_OVERLAP_THRESHOLD = 0.08            # date-token overlap
FINAL_SCORE_THRESHOLD = 0.22             # overall score to accept

# --- EXCEL COLORS ---
DARK_RED_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# --- Optional similarity ---
try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
except Exception:  # pragma: no cover
    TfidfVectorizer = None
    cosine_similarity = None


# ---------------------------
# OCR-ROBUST TEXT NORMALIZATION
# ---------------------------
def normalize_text(s: str) -> str:
    if not s:
        return ""
    s = str(s)

    # remove OCR wrappers and noise that breaks matching
    s = s.replace("<", " ").replace(">", " ")
    s = s.replace("\u00ad", "")  # soft hyphen
    s = s.replace("¬", "")       # OCR line-break marker
    s = s.replace("–", "-").replace("—", "-").replace("−", "-")

    # whitespace / hidden chars
    s = s.replace("\u00a0", " ").replace("\u200b", "").replace("\ufeff", "")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_key(s: str) -> str:
    return re.sub(r"\s+", "", normalize_text(s))


def is_empty(v) -> bool:
    return v is None or str(v).strip() == "" or str(v).strip().lower() == "nan"


def _norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


def detect_text_column(df: pd.DataFrame) -> str:
    for c in ["text", "s", "full_text", "content", "document_text", "body"]:
        if c in df.columns:
            return c
    obj_cols = [c for c in df.columns if df[c].dtype == "object"]
    if not obj_cols:
        raise ValueError("No text-like column found.")
    return max(obj_cols, key=lambda c: df[c].astype(str).str.len().mean())


# ---------------------------
# OCR-ROBUST MONTH + DATE PARSING
# ---------------------------
MONTHS = ["Januar","Februar","März","Maerz","Marz","April","Mai","Juni","Juli","August","September","Oktober","November","Dezember"]
MONTHS_MAP = {
    "januar": 1,
    "februar": 2,
    "märz": 3, "maerz": 3, "marz": 3,
    "april": 4,
    "mai": 5,
    "juni": 6,
    "juli": 7,
    "august": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "dezember": 12,
}

def _letter_spaced(word: str) -> str:
    # Allows OCR spacing/dots/hyphens between letters: "J un i", "N o v e m b e r"
    return "".join([re.escape(ch) + r"[\s\.\-]*" for ch in word])

MONTHS_RX_OCR = "(?:" + "|".join(_letter_spaced(m) for m in MONTHS) + ")"

# captures: 29. Juni 2021 (including OCR-split month)
RE_GER_DATE_OCR = re.compile(rf"\b(\d{{1,2}})\.\s*({MONTHS_RX_OCR})\s*(\d{{4}})\b", re.IGNORECASE)
RE_NUM_DATE = re.compile(r"\b(\d{1,2})\.(\d{1,2})\.(\d{4})\b")

def _cleanup_month(mstr: str) -> str:
    # keep letters only, normalize umlauts variants
    s = re.sub(r"[^A-Za-zÄÖÜäöü]", "", mstr or "").lower()
    s = s.replace("ae", "ä").replace("oe", "ö").replace("ue", "ü")
    return s

def parse_date_token(s: str) -> Optional[date]:
    s = normalize_text(s or "")
    m = RE_GER_DATE_OCR.search(s)
    if m:
        d = int(m.group(1))
        mon_key = _cleanup_month(m.group(2))
        mon = MONTHS_MAP.get(mon_key)
        y = int(m.group(3))
        if mon:
            return date(y, mon, d)
    m2 = RE_NUM_DATE.search(s)
    if m2:
        return date(int(m2.group(3)), int(m2.group(2)), int(m2.group(1)))
    return None


# ---------------------------
# VG (Verwaltungsgericht) DETECTION + AC NUMBER IN BRACKETS
# ---------------------------
# tolerate genitive and OCR noise
KW_VG_RX = re.compile(
    r"verwaltungsgericht(?:s)?\s+des\s+kantons\s+z[uü]rich",
    re.IGNORECASE,
)

RE_AC_IN_BRACKETS = re.compile(
    r"\(\s*([A-Za-zÄÖÜäöü]{1,6})\s*\.?\s*(\d{4})\s*\.?\s*(\d{4,5})\s*\)",
    re.IGNORECASE,
)

SPLIT_SENT_RX = re.compile(r"(?:[;:\n]|(?<!\d)\.)\s+")

def split_sentences(text: str) -> list[tuple[int, int, str]]:
    spans: list[tuple[int, int, str]] = []
    start = 0
    for m in SPLIT_SENT_RX.finditer(text):
        end = m.end()
        sent = text[start:end].strip()
        if sent:
            spans.append((start, end, sent))
        start = end
    tail = text[start:].strip()
    if tail:
        spans.append((start, len(text), tail))
    return spans

def find_sentence_with_kw(text: str) -> list[tuple[int, int, str]]:
    out = []
    for s_start, s_end, sent in split_sentences(text):
        if KW_VG_RX.search(sent):
            out.append((s_start, s_end, sent))
    return out


# ---------------------------
# HARD MARKERS (topic anchors) + DATE OVERLAP
# ---------------------------
RE_KAT = re.compile(r"\bKat\.\s*-?\s*Nrn?\.?\s*[A-Za-z0-9\-\/]+\b", re.IGNORECASE)
RE_ASSEK = re.compile(r"\bAssek\.\s*-?\s*Nr\.?\s*\d+\b", re.IGNORECASE)
RE_BESCHLUSSNR = re.compile(r"\bBeschluss\s+Nr\.?\s*\d+\b", re.IGNORECASE)
RE_BEHOERDE = re.compile(r"\b(Stadtrat\s+von\s+Z[uü]rich|Stadtrat\s+Z[uü]rich|Gemeinderat\s+\w+|Baudirektion\s+Kanton\s+Z[uü]rich|Bezirksrat\s+Z[uü]rich)\b", re.IGNORECASE)
RE_STREET = re.compile(r"\b([A-ZÄÖÜ][a-zäöüß]+(?:strasse|straße|weg|platz))\b", re.IGNORECASE)
RE_LOCAL = re.compile(r"\b(Z[uü]rich(?:-[A-ZÄÖÜa-zäöüß]+)?|Witikon|Wetzikon|Kanton\s+Z[uü]rich)\b", re.IGNORECASE)

def extract_hard_markers(text: str) -> set[str]:
    t = normalize_text(text)
    markers: set[str] = set()
    for rx in [RE_KAT, RE_ASSEK, RE_BESCHLUSSNR, RE_BEHOERDE, RE_STREET, RE_LOCAL]:
        for m in rx.finditer(t):
            markers.add(m.group(0).lower())
    return markers

def jaccard(a: set[str], b: set[str]) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0

# extract ALL date tokens (OCR-robust month + numeric dates)
def extract_date_tokens(text: str) -> set[str]:
    t = normalize_text(text)
    out: set[str] = set()
    for m in RE_GER_DATE_OCR.finditer(t):
        # normalize to a canonical string: dd.mm.yyyy
        d = int(m.group(1))
        mon = MONTHS_MAP.get(_cleanup_month(m.group(2)))
        y = int(m.group(3))
        if mon:
            out.add(f"{d:02d}.{mon:02d}.{y:04d}")
    for m in RE_NUM_DATE.finditer(t):
        out.add(f"{int(m.group(1)):02d}.{int(m.group(2)):02d}.{int(m.group(3)):04d}")
    return out

def date_overlap_score(a: str, b: str) -> float:
    da = extract_date_tokens(a)
    db = extract_date_tokens(b)
    if not da or not db:
        return 0.0
    inter = len(da & db)
    union = len(da | db)
    return inter / union if union else 0.0


# ---------------------------
# TFIDF SIMILARITY (optional)
# ---------------------------
GER_STOPWORDS = sorted(
    {
        "und","oder","der","die","das","den","dem","des","ein","eine","einer","eines",
        "im","in","auf","an","am","aus","bei","mit","vom","von","zu","zur","zum",
        "dass","ist","sind","war","waren","wird","werden","wurde","wurden",
        "nicht","nur","auch","als","wie","sich","es","sie","er","wir","ihr","ihre",
        "gegen","nach","vor","unter","über","ueber","uber","wegen","ohne","insbesondere",
        "daher","somit","diese","dieser","dieses","diesen","diesem",
        "entscheid","urteil","rekurs","beschwerde","verwaltungsgericht","baurekursgericht",
    }
)

def build_vectorizer():
    if TfidfVectorizer is None:
        return None
    return TfidfVectorizer(
        lowercase=True,
        stop_words=GER_STOPWORDS,
        ngram_range=(1, 2),
        min_df=1,
        max_df=0.98,
        token_pattern=r"(?u)\b[\wäöüÄÖÜß]{2,}\b",
    )

def content_similarity(a: str, b: str) -> float:
    if TfidfVectorizer is None or cosine_similarity is None:
        return 0.0
    a = normalize_text(a)[:SEARCH_FIRST_N_CHARS]
    b = normalize_text(b)[:SEARCH_FIRST_N_CHARS]
    if len(a) < 80 or len(b) < 80:
        return 0.0
    vec = build_vectorizer()
    if vec is None:
        return 0.0
    X = vec.fit_transform([a, b])
    sim = float(cosine_similarity(X[0:1], X[1:2])[0][0])
    return max(0.0, min(1.0, sim))


# ---------------------------
# FSC -> extract VG reference
# ---------------------------
@dataclass(frozen=True)
class VGRef:
    kind: str  # "number" | "date" | "none"
    ac_number: str = ""
    ac_date_str: str = ""

def extract_vg_reference_from_fsc_text(fsc_text: str) -> VGRef:
    """
    Rule:
    - Prefer AC number in brackets near a VG mention.
    - Else fall back to a date near VG mention (OCR-robust).
    """
    t = normalize_text(fsc_text)[:SEARCH_FIRST_N_CHARS]
    hits = find_sentence_with_kw(t)
    if not hits:
        return VGRef(kind="none")

    # 1) bracket-number: search in sentence + nearby window (OCR may split)
    for s_start, s_end, sent in hits:
        # try sentence first
        m = RE_AC_IN_BRACKETS.search(sent)
        if not m:
            left = max(0, s_start - NEARBY_WINDOW)
            right = min(len(t), s_end + NEARBY_WINDOW)
            m = RE_AC_IN_BRACKETS.search(t[left:right])
        if m:
            prefix, year, num = m.group(1), m.group(2), m.group(3)
            ac_num = f"{prefix}.{year}.{num}"
            return VGRef(kind="number", ac_number=ac_num)

    # 2) date near VG: search sentence + nearby window
    for s_start, s_end, sent in hits:
        m = RE_GER_DATE_OCR.search(sent)
        if not m:
            left = max(0, s_start - NEARBY_WINDOW)
            right = min(len(t), s_end + NEARBY_WINDOW)
            m = RE_GER_DATE_OCR.search(t[left:right])
        if m:
            # keep original matched substring for transparency
            return VGRef(kind="date", ac_date_str=m.group(0))

        m2 = RE_NUM_DATE.search(sent)
        if not m2:
            left = max(0, s_start - NEARBY_WINDOW)
            right = min(len(t), s_end + NEARBY_WINDOW)
            m2 = RE_NUM_DATE.search(t[left:right])
        if m2:
            return VGRef(kind="date", ac_date_str=m2.group(0))

    return VGRef(kind="none")


# ---------------------------
# MAIN
# ---------------------------
def main() -> None:
    df_raw = pd.read_excel(INPUT_XLSX, dtype=object)
    df = _norm_cols(df_raw)

    required = {"court_id", "case_number"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required column(s): {sorted(missing)}")

    text_col = detect_text_column(df)

    # ensure columns exist
    for col in [
        "first_instance", "second_instance", "third_instance",
        "fsc_vg_extracted_date", "fsc_link_status", "fsc_link_sim",
        "fsc_hard_marker_score", "fsc_date_overlap_score", "fsc_final_score",
    ]:
        if col not in df.columns:
            df[col] = ""

    # Build AC lookups
    ac_mask = df["court_id"].astype(str).str.strip().eq("AC_ZH")
    ac_keys_to_row: dict[str, int] = {}
    for i in df.index[ac_mask]:
        for col in ["case_number", "case_number_2", "case_number_full"]:
            if col in df.columns:
                val = normalize_text(df.at[i, col] or "")
                if val:
                    ac_keys_to_row[normalize_key(val)] = int(i)

    ac_date_to_rows: dict[date, list[int]] = {}
    if "decision_date" in df.columns:
        for i in df.index[ac_mask]:
            dval = parse_date_token(df.at[i, "decision_date"])
            if dval:
                ac_date_to_rows.setdefault(dval, []).append(int(i))

    # BAC lookup by number (for copying BAC first_instance across)
    bac_mask = df["court_id"].astype(str).str.strip().eq("BAC_ZH")
    bac_keys_to_row: dict[str, int] = {}
    for i in df.index[bac_mask]:
        for col in ["case_number", "case_number_2", "case_number_full"]:
            if col in df.columns:
                val = normalize_text(df.at[i, col] or "")
                if val:
                    bac_keys_to_row[normalize_key(val)] = int(i)

    # Track formatting
    changed_cells: set[tuple[int, str]] = set()
    red_cells: set[tuple[int, str]] = set()

    fsc_mask = df["court_id"].astype(str).str.strip().eq("FSC")

    # Cache AC texts + markers for faster scoring in date-only expansion
    ac_row_to_text: dict[int, str] = {}
    ac_row_to_markers: dict[int, set[str]] = {}
    for i in df.index[ac_mask]:
        t = normalize_text(df.at[i, text_col] or "")[:SEARCH_FIRST_N_CHARS]
        ac_row_to_text[int(i)] = t
        ac_row_to_markers[int(i)] = extract_hard_markers(t)

    for idx in df.index[fsc_mask]:
        fsc_case = normalize_text(df.at[idx, "case_number"] or "")
        fsc_text = normalize_text(df.at[idx, text_col] or "")[:SEARCH_FIRST_N_CHARS]
        fsc_markers = extract_hard_markers(fsc_text)

        # always set third_instance to FSC itself if empty
        if fsc_case and is_empty(df.at[idx, "third_instance"]):
            df.at[idx, "third_instance"] = fsc_case
            changed_cells.add((idx, "third_instance"))

        vgref = extract_vg_reference_from_fsc_text(fsc_text)
        if vgref.kind == "none":
            continue

        # --- case 1: AC number in brackets ---
        if vgref.kind == "number":
            ac_num = vgref.ac_number
            if ac_num:
                # set FSC second_instance
                if is_empty(df.at[idx, "second_instance"]) or normalize_key(df.at[idx, "second_instance"]) == normalize_key(ac_num):
                    if df.at[idx, "second_instance"] != ac_num:
                        df.at[idx, "second_instance"] = ac_num
                        changed_cells.add((idx, "second_instance"))

                ac_row = ac_keys_to_row.get(normalize_key(ac_num))
                if ac_row is None:
                    df.at[idx, "fsc_link_status"] = "no_ac_match_in_dataset"
                    changed_cells.add((idx, "fsc_link_status"))
                    red_cells.add((idx, "second_instance"))
                    continue

                # bidirectional FSC <-> AC
                ac_case = normalize_text(df.at[ac_row, "case_number"] or "")
                if ac_case and is_empty(df.at[ac_row, "second_instance"]):
                    df.at[ac_row, "second_instance"] = ac_case
                    changed_cells.add((ac_row, "second_instance"))
                if fsc_case and is_empty(df.at[ac_row, "third_instance"]):
                    df.at[ac_row, "third_instance"] = fsc_case
                    changed_cells.add((ac_row, "third_instance"))

                # copy BAC (from AC->FSC)
                if not is_empty(df.at[ac_row, "first_instance"]) and is_empty(df.at[idx, "first_instance"]):
                    df.at[idx, "first_instance"] = df.at[ac_row, "first_instance"]
                    changed_cells.add((idx, "first_instance"))

                # propagate to BAC row (only if empty)
                bac_cn = normalize_text(df.at[ac_row, "first_instance"] or "")
                bac_row = bac_keys_to_row.get(normalize_key(bac_cn)) if bac_cn else None
                if bac_row is not None:
                    if is_empty(df.at[bac_row, "second_instance"]) and ac_case:
                        df.at[bac_row, "second_instance"] = ac_case
                        changed_cells.add((bac_row, "second_instance"))
                    if is_empty(df.at[bac_row, "third_instance"]) and fsc_case:
                        df.at[bac_row, "third_instance"] = fsc_case
                        changed_cells.add((bac_row, "third_instance"))

                df.at[idx, "fsc_link_status"] = "ok_by_number"
                df.at[idx, "fsc_link_sim"] = 1.0
                changed_cells.add((idx, "fsc_link_status"))
                changed_cells.add((idx, "fsc_link_sim"))
            continue

        # --- case 2: date-only near VG ---
        if vgref.kind == "date":
            df.at[idx, "fsc_vg_extracted_date"] = vgref.ac_date_str
            changed_cells.add((idx, "fsc_vg_extracted_date"))

            dval = parse_date_token(vgref.ac_date_str)
            if not dval:
                df.at[idx, "fsc_link_status"] = "date_parse_failed"
                changed_cells.add((idx, "fsc_link_status"))
                red_cells.add((idx, "second_instance"))
                continue

            # candidates by decision_date first
            candidate_rows = list(ac_date_to_rows.get(dval, []))

            # if none: expand to *all* AC rows, but still require strong markers+dates
            expanded = False
            if not candidate_rows:
                candidate_rows = list(ac_row_to_text.keys())
                expanded = True

            best_row = None
            best_final = -1.0
            best_sim = 0.0
            best_hard = 0.0
            best_date = 0.0

            for r in candidate_rows:
                ac_text = ac_row_to_text.get(r, "")
                if not ac_text:
                    continue

                hard = jaccard(fsc_markers, ac_row_to_markers.get(r, set()))
                dsim = date_overlap_score(fsc_text, ac_text)
                sim = content_similarity(fsc_text, ac_text)

                # final score weights:
                # hard markers matter most for your situation (same institutions, Kat/Assek, Beschluss)
                final = 0.50 * hard + 0.30 * dsim + 0.20 * sim

                if final > best_final:
                    best_final = final
                    best_row = r
                    best_sim = sim
                    best_hard = hard
                    best_date = dsim

            df.at[idx, "fsc_hard_marker_score"] = round(best_hard, 4)
            df.at[idx, "fsc_date_overlap_score"] = round(best_date, 4)
            df.at[idx, "fsc_link_sim"] = round(best_sim, 4)
            df.at[idx, "fsc_final_score"] = round(best_final, 4)
            changed_cells.add((idx, "fsc_hard_marker_score"))
            changed_cells.add((idx, "fsc_date_overlap_score"))
            changed_cells.add((idx, "fsc_link_sim"))
            changed_cells.add((idx, "fsc_final_score"))

            # acceptance rule: avoid false positives
            ok = True
            if best_row is None:
                ok = False
            else:
                # if we expanded to all AC rows, demand stronger hard markers
                if expanded:
                    if not (best_hard >= HARD_MARKER_THRESHOLD and best_date >= DATE_OVERLAP_THRESHOLD):
                        ok = False
                # general minimums
                if best_final < FINAL_SCORE_THRESHOLD:
                    ok = False
                # also ensure at least one of: strong markers OR reasonable TFIDF
                if not (best_hard >= HARD_MARKER_THRESHOLD or best_sim >= DATE_MATCH_SIM_THRESHOLD):
                    ok = False

            if not ok:
                df.at[idx, "fsc_link_status"] = "content_mismatch_on_date"
                changed_cells.add((idx, "fsc_link_status"))
                red_cells.add((idx, "second_instance"))
                continue

            # link FSC <-> best AC
            ac_case = normalize_text(df.at[best_row, "case_number"] or "")
            if ac_case:
                if is_empty(df.at[idx, "second_instance"]) or normalize_key(df.at[idx, "second_instance"]) == normalize_key(ac_case):
                    if df.at[idx, "second_instance"] != ac_case:
                        df.at[idx, "second_instance"] = ac_case
                        changed_cells.add((idx, "second_instance"))
                if is_empty(df.at[best_row, "second_instance"]):
                    df.at[best_row, "second_instance"] = ac_case
                    changed_cells.add((best_row, "second_instance"))

            if fsc_case and is_empty(df.at[best_row, "third_instance"]):
                df.at[best_row, "third_instance"] = fsc_case
                changed_cells.add((best_row, "third_instance"))

            if fsc_case and is_empty(df.at[idx, "third_instance"]):
                df.at[idx, "third_instance"] = fsc_case
                changed_cells.add((idx, "third_instance"))

            # copy BAC from AC -> FSC if present
            if not is_empty(df.at[best_row, "first_instance"]) and is_empty(df.at[idx, "first_instance"]):
                df.at[idx, "first_instance"] = df.at[best_row, "first_instance"]
                changed_cells.add((idx, "first_instance"))

            # propagate to BAC row (only if empty)
            bac_cn = normalize_text(df.at[best_row, "first_instance"] or "")
            bac_row = bac_keys_to_row.get(normalize_key(bac_cn)) if bac_cn else None
            if bac_row is not None:
                if is_empty(df.at[bac_row, "second_instance"]) and ac_case:
                    df.at[bac_row, "second_instance"] = ac_case
                    changed_cells.add((bac_row, "second_instance"))
                if is_empty(df.at[bac_row, "third_instance"]) and fsc_case:
                    df.at[bac_row, "third_instance"] = fsc_case
                    changed_cells.add((bac_row, "third_instance"))

            df.at[idx, "fsc_link_status"] = "ok_by_date_hardmarkers"
            changed_cells.add((idx, "fsc_link_status"))

    # save
    df.to_excel(OUTPUT_XLSX, index=False)

    # formatting
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.active
    header = {str(cell.value).strip().lower(): cell.column for cell in ws[1]}

    for ridx, col in changed_cells:
        cidx = header.get(col)
        if cidx:
            ws.cell(row=ridx + 2, column=cidx).fill = LIGHT_GREEN_FILL

    for ridx, col in red_cells:
        cidx = header.get(col)
        if cidx:
            ws.cell(row=ridx + 2, column=cidx).fill = DARK_RED_FILL

    wb.save(OUTPUT_XLSX)
    print(f"[OK] Saved: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
