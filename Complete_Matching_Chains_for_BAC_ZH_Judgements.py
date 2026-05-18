from __future__ import annotations

import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

INPUT_XLSX = r"X.xlsx"
OUTPUT_XLSX = r"Y.xlsx"

SEARCH_FIRST_N_CHARS = 1500

PREFIXES = ["VB", "Art", "AEG", "AN", "Dr", "EG", "GB", "KE", "NZP", "PB", "PK", "RG", "SB", "SR", "URB", "VK", "VR"]
PREFIX_ALT = "|".join(map(re.escape, PREFIXES))

DARK_RED_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def _letter_spaced(word: str) -> str:
    return "".join([re.escape(ch) + r"[\s\.\-]*" for ch in word])


BESTATIGT_RE = _letter_spaced("bestatigt")
MIT_RE = _letter_spaced("mit")

CONFIRMED_PATTERN = re.compile(
    rf"{BESTATIGT_RE}\s*{MIT_RE}\s*(?:{PREFIX_ALT})\s*\.?\s*(\d{{4}})\s*\.?\s*(\d{{5}})",
    re.IGNORECASE | re.DOTALL,
)

CASE_INSTANCE_PATTERN = re.compile(
    rf"\b(?:{PREFIX_ALT})\s*\.?\s*(\d{{4}})\s*\.?\s*(\d{{5}})\b",
    re.IGNORECASE,
)

# "dieser bestatigt mit BG|BGr|BGE|BGer <code>"
BG_ALT = r"(?:BG|BGr|BGE|BGer)"
# Federal court-like code: <n><L>_<n..n>/<yyyy> OR <n><L>.<n..n>/<yyyy>
FSC_CODE_PATTERN = re.compile(
    rf"(?:dieser\s+)?{BESTATIGT_RE}\s*{MIT_RE}\s*{BG_ALT}\s*"
    rf"([0-9]+[A-Za-z])\s*[_\.]\s*([0-9]{{1,4}})\s*/\s*([0-9]{{4}})",
    re.IGNORECASE | re.DOTALL,
)


def detect_text_column(df: pd.DataFrame) -> str:
    for c in ["s", "text", "full_text", "content", "document_text", "body"]:
        if c in df.columns:
            return c
    obj_cols = [c for c in df.columns if df[c].dtype == "object"]
    if not obj_cols:
        raise ValueError("No text-like column found.")
    return max(obj_cols, key=lambda c: df[c].astype(str).str.len().mean())


def normalize_instance_code(year: str, num: str) -> str:
    return f"VB.{year}.{num}"


def normalize_fsc_code(prefix: str, num: str, year: str, sep: str = "_") -> str:
    prefix = prefix.strip()
    num = str(int(num))  # normalize leading zeros away
    return f"{prefix}{sep}{num}/{year}"


def is_empty(v) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s.lower() == "nan"


def pick_first_instance_value(row: pd.Series) -> str:
    v_full = str(row.get("case_number_full", "") or "").strip()
    if v_full and v_full.lower() != "nan":
        return v_full
    return str(row.get("case_number", "") or "").strip()


def extract_instance_codes(s: str) -> list[str]:
    s = (s or "").replace(" ", "")
    out: list[str] = []
    for m in CASE_INSTANCE_PATTERN.finditer(s):
        out.append(normalize_instance_code(m.group(1), m.group(2)))
    return out


def main() -> None:
    df = pd.read_excel(INPUT_XLSX)
    text_col = detect_text_column(df)

    for col in ["first_instance", "second_instance", "third_instance"]:
        if col not in df.columns:
            df[col] = ""

    # Map normalized VB.yyyy.nnnnn -> rows where case_number contains that instance code
    code_to_rows: dict[str, list[int]] = {}
    for idx, v in df["case_number"].fillna("").astype(str).items():
        for code in extract_instance_codes(v):
            code_to_rows.setdefault(code, []).append(idx)

    # Map FSC-like code -> rows where case_number equals that code (exact match after whitespace normalization)
    fsc_to_rows: dict[str, list[int]] = {}
    for idx, v in df["case_number"].fillna("").astype(str).items():
        cv = re.sub(r"\s+", "", v)
        fsc_to_rows.setdefault(cv, []).append(idx)

    rows_to_mark_red_second: list[int] = []
    rows_to_mark_red_third: list[int] = []
    cells_to_mark_green: set[tuple[int, str]] = set()

    for idx, row in df.iterrows():
        if str(row.get("court_id", "")).strip() != "BAC_ZH":
            continue
        if not is_empty(row.get("second_instance")):
            # keep existing behavior: only process rows where second_instance is empty
            continue

        full_text = str(row.get(text_col, "") or "")
        head_text = full_text[:SEARCH_FIRST_N_CHARS]

        # --- SECOND INSTANCE: only if explicit "bestatigt mit <PREFIX> yyyy.nnnnn" exists in head ---
        m2 = CONFIRMED_PATTERN.search(head_text)
        if m2:
            vb_code = normalize_instance_code(m2.group(1), m2.group(2))
            bac_first = pick_first_instance_value(row)

            if df.at[idx, "first_instance"] != bac_first:
                df.at[idx, "first_instance"] = bac_first
                cells_to_mark_green.add((idx, "first_instance"))

            if df.at[idx, "second_instance"] != vb_code:
                df.at[idx, "second_instance"] = vb_code
                cells_to_mark_green.add((idx, "second_instance"))

            match_rows = code_to_rows.get(vb_code, [])
            if match_rows:
                for r in match_rows:
                    if is_empty(df.at[r, "second_instance"]):
                        df.at[r, "second_instance"] = vb_code
                        cells_to_mark_green.add((r, "second_instance"))
                    if is_empty(df.at[r, "first_instance"]):
                        df.at[r, "first_instance"] = bac_first
                        cells_to_mark_green.add((r, "first_instance"))
            else:
                rows_to_mark_red_second.append(idx + 2)  # header row is 1

        # --- THIRD INSTANCE: only if explicit "bestatigt mit BG|BGr|BGE|BGer <code>" exists in head ---
        m3 = FSC_CODE_PATTERN.search(head_text)
        if m3:
            # Keep separator that appears in text (underscore vs dot)
            raw = m3.group(0)
            sep = "_" if "_" in raw else "."
            fsc_code = normalize_fsc_code(m3.group(1), m3.group(2), m3.group(3), sep=sep)

            if df.at[idx, "third_instance"] != fsc_code:
                df.at[idx, "third_instance"] = fsc_code
                cells_to_mark_green.add((idx, "third_instance"))

            # If this FSC code appears in case_number, copy first+second to that same row(s)
            key = re.sub(r"\s+", "", fsc_code)
            match_rows = fsc_to_rows.get(key, [])
            if match_rows:
                for r in match_rows:
                    if is_empty(df.at[r, "third_instance"]):
                        df.at[r, "third_instance"] = fsc_code
                        cells_to_mark_green.add((r, "third_instance"))
                    if not is_empty(df.at[idx, "first_instance"]) and is_empty(df.at[r, "first_instance"]):
                        df.at[r, "first_instance"] = df.at[idx, "first_instance"]
                        cells_to_mark_green.add((r, "first_instance"))
                    if not is_empty(df.at[idx, "second_instance"]) and is_empty(df.at[r, "second_instance"]):
                        df.at[r, "second_instance"] = df.at[idx, "second_instance"]
                        cells_to_mark_green.add((r, "second_instance"))
            else:
                rows_to_mark_red_third.append(idx + 2)

    df.to_excel(OUTPUT_XLSX, index=False)

    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.active
    header = {cell.value: cell.column for cell in ws[1]}

    # Green fill for changed cells
    for df_row_idx, col_name in cells_to_mark_green:
        col_idx = header.get(col_name)
        if not col_idx:
            continue
        ws.cell(row=df_row_idx + 2, column=col_idx).fill = LIGHT_GREEN_FILL

    # Red fills (override green)
    sec_col = header.get("second_instance")
    third_col = header.get("third_instance")
    if sec_col:
        for excel_row in rows_to_mark_red_second:
            ws.cell(row=excel_row, column=sec_col).fill = DARK_RED_FILL
    if third_col:
        for excel_row in rows_to_mark_red_third:
            ws.cell(row=excel_row, column=third_col).fill = DARK_RED_FILL

    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
