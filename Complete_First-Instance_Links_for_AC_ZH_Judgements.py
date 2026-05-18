from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from datetime import date
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

INPUT_XLSX = r"X.xlsx"
OUTPUT_XLSX = r"Y.xlsx"

SEARCH_FIRST_N_CHARS = 20000
NEARBY_WINDOW = 300  # chars before/after match where Baurekursgericht mention is allowed

# NEW: minimum similarity required to auto-link a unique date match
SIM_THRESHOLD = 28  # 0..100; tune if needed

DARK_RED_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

CUTOFF_RENAME = date(2011, 1, 1)

MONTHS_RX = r"(?:Januar|Februar|März|Maerz|Marz|April|Mai|Juni|Juli|August|September|Oktober|November|Dezember)"
DATE_TOKEN_RX = rf"(?:\d{{1,2}}\.\s*{MONTHS_RX}\s+\d{{4}}|\d{{1,2}}\.\d{{1,2}}\.\d{{4}})"

MONTHS_MAP = {
    "januar": 1,
    "februar": 2,
    "märz": 3,
    "maerz": 3,
    "marz": 3,
    "april": 4,
    "mai": 5,
    "juni": 6,
    "juli": 7,
    "august": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "dezember": 12,
}

RE_GER_DATE = re.compile(rf"\b(\d{{1,2}})\.\s*({MONTHS_RX})\s+(\d{{4}})\b", re.IGNORECASE)
RE_NUM_DATE = re.compile(r"\b(\d{1,2})\.(\d{1,2})\.(\d{4})\b", re.IGNORECASE)

KW_BAC_RX = re.compile(
    r"(Baurekursgericht|Baurekurskommission(?:\s+[IVX]{1,4})?|Baurekurskommission\s*III|Rekurskommission\s*III)",
    re.IGNORECASE,
)

# Judgment number formats like: "BRGE III Nr. 0093/2019", "III Nr. 35/2019"
RE_LOWER_NO = re.compile(
    r"\b(?:(BRGE|BRK|BKR|BRKE)\s*)?"
    r"([IVX]{1,4})\s*Nrn?\.?\s*0?(\d{1,4})/(\d{4})"
    r"(?:\s*(?:und|&|,|–|-)\s*0?(\d{1,4})/(\d{4}))?",
    re.IGNORECASE,
)

DATE_PHRASE_RES: list[re.Pattern] = [
    re.compile(rf"\bMit\s+(?:Rekursentscheid\s+)?Entscheid\s+vom\s+(?P<date>{DATE_TOKEN_RX})\b", re.IGNORECASE),
    re.compile(rf"\bmit\s+(?:Rekursentscheid|Entscheid)\s+vom\s+(?P<date>{DATE_TOKEN_RX})\b", re.IGNORECASE),
    re.compile(rf"\bRekursentscheid\s+vom\s+(?P<date>{DATE_TOKEN_RX})\b", re.IGNORECASE),
    re.compile(rf"\bEntscheid\s*(?:\([^)]{{0,120}}\))?\s*vom\s+(?P<date>{DATE_TOKEN_RX})\b", re.IGNORECASE),
    re.compile(rf"\bAm\s+(?P<date>{DATE_TOKEN_RX})\b(?P<tail>.{{0,220}}?)\bentschied\b", re.IGNORECASE),
    re.compile(rf"\bentschied\b(?P<tail>.{{0,240}}?)\bam\s+(?P<date>{DATE_TOKEN_RX})\b", re.IGNORECASE),
    re.compile(rf"\bhiess\b(?P<tail>.{{0,240}}?)\bam\s+(?P<date>{DATE_TOKEN_RX})\b", re.IGNORECASE),
    re.compile(rf"\bwies\b(?P<tail>.{{0,240}}?)\bam\s+(?P<date>{DATE_TOKEN_RX})\b", re.IGNORECASE),
]


@dataclass(frozen=True)
class Candidate:
    kind: str  # "number" | "date"
    value: str
    score: float
    evidence: str
    note: str = ""


def _norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df


def detect_text_column(df: pd.DataFrame) -> str:
    for c in ["text", "s", "full_text", "content", "document_text", "body"]:
        if c in df.columns:
            return c
    obj_cols = [c for c in df.columns if df[c].dtype == "object"]
    if not obj_cols:
        raise ValueError("No text-like column found.")
    return max(obj_cols, key=lambda c: df[c].astype(str).str.len().mean())


def is_empty(v) -> bool:
    return v is None or str(v).strip() == "" or str(v).strip().lower() == "nan"


def normalize_text(s: str) -> str:
    if not s:
        return ""
    s = str(s)
    s = s.replace("<", " ").replace(">", " ")
    s = s.replace("\u00ad", "")
    s = s.replace("¬", "")
    s = s.replace("–", "-").replace("—", "-").replace("−", "-")
    s = s.replace("\u00a0", " ").replace("\u200b", "").replace("\ufeff", "")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def parse_date_token(s: str) -> Optional[date]:
    m = RE_GER_DATE.search(s or "")
    if m:
        d = int(m.group(1))
        mon = MONTHS_MAP[m.group(2).lower()]
        y = int(m.group(3))
        return date(y, mon, d)
    m = RE_NUM_DATE.search(s or "")
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return None


def normalize_lower_number(m: re.Match) -> tuple[str, str]:
    roman = m.group(2).upper()
    n1 = int(m.group(3))
    y1 = int(m.group(4))
    primary = f"{roman} Nr. {n1:04d}/{y1}"

    note = ""
    if m.group(5) and m.group(6):
        n2 = int(m.group(5))
        y2 = int(m.group(6))
        note = f"multi_numbers={primary} und {roman} Nr. {n2:04d}/{y2}"
    return primary, note


def split_sentences(text: str) -> list[tuple[int, int, str]]:
    spans: list[tuple[int, int, str]] = []
    start = 0
    for m in re.finditer(r"(?:[;:\n]|(?<!\d)\.)\s+", text):
        end = m.end()
        sent = text[start:end].strip()
        if sent:
            spans.append((start, end, sent))
        start = end
    tail = text[start:].strip()
    if tail:
        spans.append((start, len(text), tail))
    return spans


def evidence_window(text: str, start: int, end: int, pad: int = 220) -> str:
    left = max(0, start - pad)
    right = min(len(text), end + pad)
    window = text[left:right]
    window = re.sub(r"\s+", " ", window.replace("\n", " ")).strip()
    return window[:700]


def has_kw_near(text: str, match_start: int, match_end: int) -> bool:
    left = max(0, match_start - NEARBY_WINDOW)
    right = min(len(text), match_end + NEARBY_WINDOW)
    return bool(KW_BAC_RX.search(text[left:right]))


def date_phrase_score(dval: date, phrase_strength: float, kw_near: bool, sentence_has_kw: bool) -> float:
    score = phrase_strength
    if kw_near or sentence_has_kw:
        score += 0.03
    if dval >= CUTOFF_RENAME:
        pass
    return min(0.99, max(0.0, score))


def best_lower_court_reference(text: str) -> Optional[Candidate]:
    text = normalize_text(text)
    sentences = split_sentences(text)
    candidates: list[Candidate] = []

    for s_start, _, sent in sentences:
        sent_has_kw = bool(KW_BAC_RX.search(sent))
        for m in RE_LOWER_NO.finditer(sent):
            val, note = normalize_lower_number(m)
            abs_start = s_start + m.start()
            abs_end = s_start + m.end()
            if not (sent_has_kw or has_kw_near(text, abs_start, abs_end)):
                continue
            candidates.append(
                Candidate(
                    kind="number",
                    value=val,
                    score=0.95,
                    evidence=evidence_window(text, abs_start, abs_end),
                    note=note,
                )
            )

    for s_start, _, sent in sentences:
        sent_has_kw = bool(KW_BAC_RX.search(sent))
        for rx in DATE_PHRASE_RES:
            for m in rx.finditer(sent):
                dstr = m.group("date")
                dval = parse_date_token(dstr)
                if not dval:
                    continue
                abs_start = s_start + m.start()
                abs_end = s_start + m.end()
                kw_near = has_kw_near(text, abs_start, abs_end) or sent_has_kw
                if not kw_near:
                    continue

                pattern = rx.pattern.lower()
                if "rekursentscheid" in pattern:
                    base = 0.97
                elif "mit" in pattern and "entscheid" in pattern and "vom" in pattern:
                    base = 0.97
                elif "entscheid" in pattern and "vom" in pattern:
                    base = 0.96
                elif "am" in pattern and "entschied" in pattern:
                    base = 0.95
                elif "entschied" in pattern and "am" in pattern:
                    base = 0.95
                elif "hiess" in pattern or "wies" in pattern:
                    base = 0.92
                else:
                    base = 0.90

                score = date_phrase_score(dval, base, kw_near=True, sentence_has_kw=sent_has_kw)
                candidates.append(
                    Candidate(
                        kind="date",
                        value=dstr,
                        score=score,
                        evidence=evidence_window(text, abs_start, abs_end),
                        note="",
                    )
                )

    if not candidates:
        return None
    candidates.sort(key=lambda c: (c.score, c.kind == "number"), reverse=True)
    return candidates[0]


# ---------- NEW: similarity signals (keywords + other dates + identifiers) ----------

GER_STOPWORDS = {
    "der", "die", "das", "und", "oder", "in", "im", "am", "an", "auf", "aus", "zu", "zur", "zum", "von", "vom",
    "mit", "ohne", "gegen", "für", "fur", "bei", "nach", "vor", "über", "uber", "unter", "zwischen", "wird", "werden",
    "ist", "sind", "war", "waren", "sei", "hat", "haben", "wurde", "würde", "kann", "können", "nicht", "nur", "auch",
    "als", "dass", "dem", "den", "des", "ein", "eine", "einer", "eines", "einem", "einen", "dies", "diese", "dieser",
    "dieses", "sowie", "insbesondere", "betreffend", "hinsichtlich", "gemäss", "gemaess", "unter", "unter anderem",
    "art", "abs", "lit", "ziff", "nr", "nrn", "nrs", "bge", "bger", "urteil", "entscheid", "beschluss", "kammer",
    "abteilung", "kanton", "kantons", "zürich", "zurich",
}
TOKEN_RX = re.compile(r"[A-Za-zÄÖÜäöüß]+", re.UNICODE)

RE_KAT = re.compile(r"\b(?:kat\.?-?\s*nrn?\.?|kataster(?:nummer)?|ktn|parzelle|grundstück)\s*[:\-]?\s*(\d{1,6})\b", re.IGNORECASE)
RE_GEMEINDE = re.compile(r"\bGemeinde\s+([A-ZÄÖÜ][A-Za-zÄÖÜäöüß\-]{2,})\b", re.UNICODE)

def extract_keywords(text: str, *, top_k: int = 12) -> list[str]:
    text = normalize_text(text or "").lower()
    toks = [t for t in TOKEN_RX.findall(text) if len(t) >= 4]
    toks = [t for t in toks if t not in GER_STOPWORDS]
    if not toks:
        return []
    counts = Counter(toks)
    ranked = sorted(counts.items(), key=lambda x: (-x[1], x[0]))
    return [w for w, _ in ranked[:top_k]]

def extract_all_dates(text: str) -> set[date]:
    """
    All dates mentioned (submission, publication, decisions, etc.).
    """
    t = normalize_text(text or "")
    out: set[date] = set()
    for m in RE_GER_DATE.finditer(t):
        d = int(m.group(1))
        mon = MONTHS_MAP[m.group(2).lower()]
        y = int(m.group(3))
        try:
            out.add(date(y, mon, d))
        except ValueError:
            pass
    for m in RE_NUM_DATE.finditer(t):
        try:
            out.add(date(int(m.group(3)), int(m.group(2)), int(m.group(1))))
        except ValueError:
            pass
    return out

def extract_identifiers(text: str) -> dict[str, set[str]]:
    t = normalize_text(text or "")
    kats = set(RE_KAT.findall(t))
    gemeinden = set([m.group(1) for m in RE_GEMEINDE.finditer(t)])
    gemeinden = set([g.strip() for g in gemeinden if g.strip()])
    return {"kat": kats, "gemeinde": gemeinden}

def jaccard(a: set[str], b: set[str]) -> float:
    if not a and not b:
        return 0.0
    inter = a & b
    union = a | b
    return (len(inter) / len(union)) if union else 0.0

def sim_score_bundle(
    ac_text: str,
    bac_text: str,
    *,
    ac_keywords: list[str],
    bac_keywords: list[str],
) -> tuple[int, str, bool]:
    """
    Returns (score 0..100, debug_string, strong_id_match_bool)
    """
    ac_ids = extract_identifiers(ac_text)
    bac_ids = extract_identifiers(bac_text)

    # Strong signals
    kat_overlap = ac_ids["kat"] & bac_ids["kat"]
    strong_id = len(kat_overlap) > 0

    gemeinde_overlap = ac_ids["gemeinde"] & bac_ids["gemeinde"]

    # Softer signals
    ac_dates = extract_all_dates(ac_text)
    bac_dates = extract_all_dates(bac_text)
    date_overlap = ac_dates & bac_dates

    kw_overlap = set(ac_keywords) & set(bac_keywords)

    # Scores
    kw_sim = jaccard(set(ac_keywords), set(bac_keywords))          # 0..1
    date_sim = jaccard(set(map(str, ac_dates)), set(map(str, bac_dates)))  # 0..1
    gemeinde_sim = 1.0 if gemeinde_overlap else 0.0
    kat_sim = 1.0 if strong_id else 0.0

    # Weighted total (tuned for legal texts: IDs/dates are strong disambiguators)
    # - kat/parcel match is extremely strong
    # - municipality match helps
    # - other-date overlap reduces false positives when same decision_date is shared
    # - keyword overlap adds topical alignment
    total = (
        0.45 * kat_sim +
        0.15 * gemeinde_sim +
        0.20 * date_sim +
        0.20 * kw_sim
    )

    score = int(round(total * 100))
    dbg = (
        f"score={score} "
        f"kat_match={sorted(kat_overlap)[:3]} "
        f"gemeinde_match={sorted(gemeinde_overlap)[:3]} "
        f"date_overlap_count={len(date_overlap)} "
        f"kw_overlap={sorted(list(kw_overlap))[:6]}"
    )
    return score, dbg, strong_id


def main() -> None:
    df_raw = pd.read_excel(INPUT_XLSX, dtype=object)
    df = _norm_cols(df_raw)

    required = {"court_id", "case_number"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required column(s): {sorted(missing)}")

    text_col = detect_text_column(df)
    print(f"[INFO] Using text column: {text_col}")

    for col in [
        "first_instance",
        "second_instance",
        "first_instance_date",
        "first_instance_conf",
        "first_instance_note",
        "first_instance_evidence",
        "first_instance_date_match_status",
        "first_instance_date_match_cases",
        "first_instance_date_match_bac_keywords",
        "ac_keywords",
        # NEW
        "sim_score",
        "sim_debug",
    ]:
        if col not in df.columns:
            df[col] = ""

    bac_mask = df["court_id"].astype(str).str.strip().eq("BAC_ZH")
    ac_mask = df["court_id"].astype(str).str.strip().eq("AC_ZH")

    # BAC decision_date -> case_numbers + row indices
    bac_date_to_cases: dict[date, list[str]] = {}
    bac_date_to_rows: dict[date, list[int]] = {}
    if "decision_date" in df.columns:
        for i in df.index[bac_mask]:
            dval = parse_date_token(normalize_text(df.at[i, "decision_date"] or ""))
            if not dval:
                continue
            cn = normalize_text(df.at[i, "case_number"] or "")
            if cn:
                bac_date_to_cases.setdefault(dval, []).append(cn)
            bac_date_to_rows.setdefault(dval, []).append(int(i))

        # de-dupe preserve order
        for k, v in list(bac_date_to_cases.items()):
            seen: set[str] = set()
            out: list[str] = []
            for item in v:
                if item not in seen:
                    seen.add(item)
                    out.append(item)
            bac_date_to_cases[k] = out

    def date_match_cases(dstr: str) -> list[str]:
        dval = parse_date_token(normalize_text(dstr))
        if not dval:
            return []
        return bac_date_to_cases.get(dval, [])

    def date_match_rows(dstr: str) -> list[int]:
        dval = parse_date_token(normalize_text(dstr))
        if not dval:
            return []
        return bac_date_to_rows.get(dval, [])

    # IMPORTANT: only process AC rows where BOTH first_instance and second_instance are empty
    empty_first_mask = df["first_instance"].apply(is_empty)
    empty_second_mask = df["second_instance"].apply(is_empty)
    target_mask = ac_mask & empty_first_mask & empty_second_mask

    print(f"[INFO] AC_ZH rows: {int(ac_mask.sum())}")
    print(f"[INFO] AC_ZH with empty first_instance AND second_instance: {int(target_mask.sum())}")

    # Precompute BAC text, keywords for similarity checks
    bac_row_to_text: dict[int, str] = {}
    bac_row_to_keywords: dict[int, list[str]] = {}
    for i in df.index[bac_mask]:
        txt = str(df.at[i, text_col] or "")[:SEARCH_FIRST_N_CHARS]
        bac_row_to_text[int(i)] = normalize_text(txt)
        bac_row_to_keywords[int(i)] = extract_keywords(txt, top_k=12)

    changed_cells: set[tuple[int, str]] = set()
    red_cells: set[tuple[int, str]] = set()

    for idx in df.index[target_mask]:
        raw_txt = str(df.at[idx, text_col] or "")[:SEARCH_FIRST_N_CHARS]
        cand = best_lower_court_reference(raw_txt)
        if not cand:
            continue

        ac_text_norm = normalize_text(raw_txt)
        ac_case_number = normalize_text(df.at[idx, "case_number"] or "")

        # AC keywords (for manual compare)
        ac_kws = extract_keywords(ac_text_norm, top_k=12)
        ac_kw_str = ", ".join(ac_kws)
        if ac_kw_str and df.at[idx, "ac_keywords"] != ac_kw_str:
            df.at[idx, "ac_keywords"] = ac_kw_str
            changed_cells.add((idx, "ac_keywords"))

        # Only safe because target_mask ensures second_instance is empty
        if ac_case_number and df.at[idx, "second_instance"] != ac_case_number:
            df.at[idx, "second_instance"] = ac_case_number
            changed_cells.add((idx, "second_instance"))

        conf = int(round(cand.score * 100))
        if df.at[idx, "first_instance_conf"] != conf:
            df.at[idx, "first_instance_conf"] = conf
            changed_cells.add((idx, "first_instance_conf"))
        if df.at[idx, "first_instance_note"] != cand.note:
            df.at[idx, "first_instance_note"] = cand.note
            changed_cells.add((idx, "first_instance_note"))
        if df.at[idx, "first_instance_evidence"] != cand.evidence:
            df.at[idx, "first_instance_evidence"] = cand.evidence
            changed_cells.add((idx, "first_instance_evidence"))

        for col in ["first_instance_date_match_status", "first_instance_date_match_cases", "first_instance_date_match_bac_keywords", "sim_score", "sim_debug"]:
            if not is_empty(df.at[idx, col]):
                df.at[idx, col] = ""
                changed_cells.add((idx, col))

        if cand.kind == "number":
            # keep old behavior for explicit numbers
            if df.at[idx, "first_instance"] != cand.value:
                df.at[idx, "first_instance"] = cand.value
                changed_cells.add((idx, "first_instance"))
            if not is_empty(df.at[idx, "first_instance_date"]):
                df.at[idx, "first_instance_date"] = ""
                changed_cells.add((idx, "first_instance_date"))
            # still mark if no BAC match exists elsewhere (same as before)
            # (left unchanged intentionally)
            continue

        # Date-only candidate
        if df.at[idx, "first_instance_date"] != cand.value:
            df.at[idx, "first_instance_date"] = cand.value
            changed_cells.add((idx, "first_instance_date"))

        matches = date_match_cases(cand.value)
        match_rows = date_match_rows(cand.value)

        if not matches:
            if df.at[idx, "first_instance_date_match_status"] != "no_match":
                df.at[idx, "first_instance_date_match_status"] = "no_match"
                changed_cells.add((idx, "first_instance_date_match_status"))
            red_cells.add((idx, "first_instance_date"))
            continue

        status = "ok" if len(matches) == 1 else "ambiguous"
        joined = ", ".join(matches)

        if df.at[idx, "first_instance_date_match_status"] != status:
            df.at[idx, "first_instance_date_match_status"] = status
            changed_cells.add((idx, "first_instance_date_match_status"))
        if df.at[idx, "first_instance_date_match_cases"] != joined:
            df.at[idx, "first_instance_date_match_cases"] = joined
            changed_cells.add((idx, "first_instance_date_match_cases"))

        # If ambiguous: store BAC keywords for manual review (same as before)
        if status == "ambiguous" and match_rows:
            parts: list[str] = []
            for r in match_rows:
                bac_cn = normalize_text(df.at[r, "case_number"] or "")
                kws = bac_row_to_keywords.get(int(r), [])
                parts.append(f"{bac_cn}: " + ", ".join(kws))
            kw_blob = " | ".join([p for p in parts if p.strip()])
            if kw_blob and df.at[idx, "first_instance_date_match_bac_keywords"] != kw_blob:
                df.at[idx, "first_instance_date_match_bac_keywords"] = kw_blob
                changed_cells.add((idx, "first_instance_date_match_bac_keywords"))
            continue

        # Unique BAC-by-date: NEW -> require similarity gate before linking
        bac_row = int(match_rows[0]) if match_rows else None
        bac_text = bac_row_to_text.get(bac_row, "") if bac_row is not None else ""
        bac_kws = bac_row_to_keywords.get(bac_row, [])

        s_score, s_dbg, strong_id = sim_score_bundle(
            ac_text_norm,
            bac_text,
            ac_keywords=ac_kws,
            bac_keywords=bac_kws,
        )

        df.at[idx, "sim_score"] = s_score
        df.at[idx, "sim_debug"] = s_dbg
        changed_cells.add((idx, "sim_score"))
        changed_cells.add((idx, "sim_debug"))

        # Only link if similarity passes or strong ID matches (Kat/KTN etc.)
        if not (strong_id or s_score >= SIM_THRESHOLD):
            # Do NOT link; keep first_instance empty
            if df.at[idx, "first_instance_date_match_status"] != "content_mismatch":
                df.at[idx, "first_instance_date_match_status"] = "content_mismatch"
                changed_cells.add((idx, "first_instance_date_match_status"))
            continue

        # Link: AC row first_instance = BAC case_number (unique)
        bac_case_number = normalize_text(matches[0])
        if bac_case_number and df.at[idx, "first_instance"] != bac_case_number:
            df.at[idx, "first_instance"] = bac_case_number
            changed_cells.add((idx, "first_instance"))

        # BAC side linkage: NEVER overwrite second_instance if filled
        if bac_row is not None:
            bac_self = normalize_text(df.at[bac_row, "case_number"] or "")
            if bac_self and is_empty(df.at[bac_row, "first_instance"]):
                df.at[bac_row, "first_instance"] = bac_self
                changed_cells.add((bac_row, "first_instance"))
            if ac_case_number and is_empty(df.at[bac_row, "second_instance"]):
                df.at[bac_row, "second_instance"] = ac_case_number
                changed_cells.add((bac_row, "second_instance"))

    df.to_excel(OUTPUT_XLSX, index=False)

    wb = load_workbook(OUTPUT_XLSX)
    ws = wb.active
    header = {str(cell.value).strip().lower(): cell.column for cell in ws[1]}

    for ridx, col in changed_cells:
        cidx = header.get(col)
        if cidx:
            ws.cell(row=ridx + 2, column=cidx).fill = LIGHT_GREEN_FILL

    for ridx, col in red_cells:
        cidx = header.get(col)
        if cidx:
            ws.cell(row=ridx + 2, column=cidx).fill = DARK_RED_FILL

    wb.save(OUTPUT_XLSX)
    print(f"[OK] Saved output: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
